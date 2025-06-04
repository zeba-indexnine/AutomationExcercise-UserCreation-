from openpyxl import load_workbook
from robot.api.deco import keyword

@keyword
def fetch_testdata_by_id(file_path, target_id):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict

        raise ValueError(f"Target ID '{target_id}' not found in the Excel file.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")


